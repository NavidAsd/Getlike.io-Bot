# windows ver 1.0
from logging import exception
from urllib import request
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from subprocess import call
import pickle
import playsound
import platform
import webbrowser
import pyautogui
import datetime
import subprocess
#import Connection_CL
import pyfiglet
import openpyxl 
import random
import math
import time
import csv
import os 


noow=datetime.datetime.now()
Icon='./Files/Icon/I2.ico'

GetlikeUrl='https://getlike.io/login'
FolowUrl='https://getlike.io/tasks/instagram/subscribe/'
LikeUrl='https://getlike.io/tasks/instagram/like/'
InstagramUrl='https://www.instagram.com/accounts/login/'

Banner_font={0:'5lineoblique',1:'arrows' ,2:'block' ,3:'banner3-D' ,4:'bell' ,5:'binary' ,6:'kban' ,7:'contrast' ,8:'cosmike' ,9:'cosmic' ,10:'isometric2' ,11:'cosmic' ,12:'nancyj-underlined' ,13:'sblood'}

class Program():
    def __init__(self,User,Pass,width,height,count,FolowCount,LikeCount,Folowed,Liked,AccountCount,ccv,InstaUser,InstaPass,RepeatFolowTask,RepeatLikeTask,alarmcount,Choice,Code,firsttime,LsTime,CheckTime,result,FollowLimit,LikeLimit,FollowAccess,LikeAccess,FollowLimitInDay,LikeLimitInDay,LimitDate,NewTaskId,MoreTaskFaileCount,TaskFaileCount,UseChrome):
        self.ccv=ccv
        self.User=User
        self.Pass=Pass
        self.Code=Code
        self.result=result
        self.LsTime=LsTime
        self.Choice=Choice
        self.Liked=Liked
        self.count=count
        self.width=width
        self.CheckTime=CheckTime
        self.height=height
        self.Folowed=Folowed
        self.firsttime=firsttime
        self.LikeCount=LikeCount
        self.InstaUser=InstaUser
        self.alarmcount=alarmcount
        self.RepeatLikeTask=RepeatLikeTask
        self.RepeatFolowTask=RepeatFolowTask
        self.InstaPass=InstaPass
        self.FolowCount=FolowCount
        self.AccountCount=AccountCount
        self.FollowLimit=FollowLimit
        self.LikeLimit=LikeLimit
        self.FolloAccess=FollowAccess
        self.LikeAccess=LikeAccess
        self.FollowLimitInDay=FollowLimitInDay
        self.LikeLimitInDay=LikeLimitInDay
        self.LimitDate=LimitDate
        self.NewTaskId=NewTaskId
        self.MoreTaskFaileCount=MoreTaskFaileCount
        self.TaskFaileCount=TaskFaileCount
        self.UseChrome=UseChrome

    
    #region ObjectFinder       
    def WaitForObject(self,type,string):
        return WebDriverWait(self.driver,9).until(EC.presence_of_element_located((type,string)))

    def WaitForObjects(self,type,string):
        return WebDriverWait(self.driver,4).until(EC.presence_of_all_elements_located((type,string))) 
    #endregion

    def Explain_Tsituation(self,Text,Color,Url,State):
        print()
        if not State:
            print(Color+Text+'\33[37m')
            inp=input('Open the relevant page in the browser (y/n)? ')
            if inp =='y' or inp =='Y' or inp =='yes' or inp=='Yes':
                webbrowser.open(Url,new=1)
        else:
            print(Color+Text+'\33[37m')
      
    
    def Initial_Configuration(self):
        self.InstaUser=self.InstaPass=self.Pass=self.User=''
        os.system('mode 70,60')
        resolution = pyautogui.size()
        resault=str(resolution).replace('Size','').replace('(','').replace(')','').replace('width=','').replace(', height=','')
        if(resault.__len__()==7 or resault.__len__()==8):
            self.width=int(resault[:4])
            self.height=int(resault[4:])
            self.width=(self.width/2)+15
            self.height-=30
        elif(resault.__len__()==6):
            self.width=int(resault[:3])
            self.height=int(resault[3:])
            self.width=(self.width/2)+20
            self.height-=30
        Bot.Check_Connection()
        
    def Check_Connection(self):
        try:
            Bot.WriteLog()
        except:
            print('\33[31m'+'Connection Lost!'+'\33[37m')
            print('\33[32m'+'Check your network connection and try again'+'\33[37m')
            Bot.Exit()    

    def WriteLog(self):
        file=open("./Files/log.txt","w+")
        file.write(str(datetime.datetime.now()))
        file.write('\n')
        file.write(f'os type: {platform.system()}')
        file.write('\n')
        file.write('data storage type: TXT(file)')
        file.write('\n')
        file.write('--------------------------------')
        file.close()
        self.count=1
        Bot.GetConfig()
        
    def GetConfig(self):
        try:
            file = open('./Files/Configuration/Config.txt','r')
            data = csv.reader(file,delimiter=':')
        except:
            print('\33[31m'+'"Config.txt" file not found in "./Files/Configuration/" path!'+'\33[37m')
            fl=open('./Files/Configuration/Config.txt','w')
            fl.write('Follow:Yes')
            fl.write('\n')
            fl.write('Like:Yes')
            fl.write('\n')
            fl.write('LikeCount:100')
            fl.write('\n')
            fl.write('FollowCount:50')
            fl.write('\n')
            fl.write('UseChrome:Yes')
            fl.close()
            print('\33[32m'+'After entering the data and closing the Notepad, you can use the program.'+'\33[37m')
            subprocess.run(['notepad.exe', './Files/Configuration/Config.txt'])
            Bot.GetConfig()
        
        cv = 0
        for i in data:
            cv+=1
            if(cv==1):
                if(("".join(i[1:2]).strip()).lower() == 'yes'):
                    self.FolloAccess = True
                else:
                    self.FolloAccess = False
            elif(cv==2):
                if(("".join(i[1:2]).strip()).lower() == 'yes'):
                    self.LikeAccess = True
                else:
                    self.LikeAccess = False
            elif(cv==3):
                self.LikeLimit =("".join(i[1:2]).strip())
            elif (cv==4):
                self.FollowLimit = ("".join(i[1:2]).strip())
            elif(cv==5):
                if(("".join(i[1:2]).strip()).lower() == 'yes'):
                    self.UseChrome = True
                else:
                    self.UseChrome = False

        
        if(int(self.FollowLimit) > 0 and int(self.LikeLimit) > 0):
            self.FollowLimit = int(self.FollowLimit)
            self.LikeLimit = int(self.LikeLimit)
            Bot.Get_AllAcountCount()
        else:
            print('\33[31m'+'The "Config.txt" file data is incomplete'+'\33[37m')
            print('\33[32m'+'If you have forgotten the correct format, delete the text file and run the program again'+'\33[37m')
            subprocess.run(['notepad.exe', './Files/Configuration/Config.txt'])
            Bot.GetConfig()

    def GetStatistics(self):
        try:
            workbook =openpyxl.load_workbook('./Files/Configuration/Accounts-Statistics.xlsx')
            sheet = workbook.active
        except Exception as ex:
            Bot.WriteStatistics(1,1)
            Bot.GetStatistics()

        cv = False
        for i in range(1,sheet.max_row+1):
            if(sheet[f'A{i}'].value == self.InstaUser):
                self.FollowLimitInDay = sheet[f'B{i}'].value
                self.LikeLimitInDay = sheet[f'C{i}'].value
                self.LimitDate = sheet[f'D{i}'].value
                cv=True
                break

        if(cv == False):
            Bot.WriteStatistics(1,1)

        Bot.Start()
        
    def WriteStatistics(self,FollowLimit,LikeLimit):
        try:
            workbook =openpyxl.load_workbook('./Files/Configuration/Accounts-Statistics.xlsx')
            sheet = workbook.active
        except:
            wd=openpyxl.Workbook()
            sh = wd.active
            sh[f'A{sh.max_row}'] = "InstaUser"
            sh[f'B{sh.max_row}'] = 'FollowLimit'
            sh[f'C{sh.max_row}'] = 'LikeLimit'
            sh[f'D{sh.max_row}'] = 'LimitDate'
            wd.save('./Files/Configuration/Accounts-Statistics.xlsx')
            Bot.WriteStatistics(FollowLimit,LikeLimit)

        cv=False
        for i in range(1, sheet.max_row+1):
            if(sheet[f'A{i}'].value == self.InstaUser):
                sheet[f'B{i}'] = FollowLimit
                sheet[f'C{i}'] = LikeLimit
                if(FollowLimit <= 1 and LikeLimit <= 1):
                    sheet[f'D{sheet.max_row}'] = datetime.datetime.now()
                self.LimitDate = sheet[f'D{sheet.max_row}'].value
                cv=True
                break
        if(cv == False):
            sheet[f'A{sheet.max_row+1}'] = self.InstaUser
            sheet[f'B{sheet.max_row}'] = FollowLimit
            sheet[f'C{sheet.max_row}'] = LikeLimit
            sheet[f'D{sheet.max_row}'] = datetime.datetime.now()
            self.LimitDate = datetime.datetime.now()

        self.LikeLimitInDay = LikeLimit
        self.FollowLimitInDay = FollowLimit
        try:
            workbook.save('./Files/Configuration/Accounts-Statistics.xlsx')
        except:
            print('\33[32m'+'Please close "Accounts-Statistics.xlsx" File')
            input('Press any key to Continue..'+'\33[37m')
            Bot.WriteStatistics(FollowLimit,LikeLimit)

    def PlusLimitStatistics(self,FollowPlus,LikePlus):
        try:
            workbook =openpyxl.load_workbook('./Files/Configuration/Accounts-Statistics.xlsx')
            sheet = workbook.active
        except:
            Bot.WriteStatistics(1,1)
            
        for i in range(1, sheet.max_row+1):
            if(sheet[f'A{i}'].value == self.InstaUser):
                if(FollowPlus > 0):
                    self.FollowLimitInDay = sheet[f'B{i}'].value + FollowPlus
                    sheet[f'B{i}'] = sheet[f'B{i}'].value + FollowPlus
                if(LikePlus > 0):
                    sheet[f'C{i}'] = sheet[f'C{i}'].value + LikePlus 
                    self.LikeLimitInDay = sheet[f'C{i}'].value + LikePlus
                break

        try:
            workbook.save('./Files/Configuration/Accounts-Statistics.xlsx')
        except:
            print('\33[32m'+'Please close "Accounts-Statistics.xlsx" File')
            input('Press any key to Continue..'+'\33[37m')
            Bot.PlusLimitStatistics(FollowPlus,LikePlus)

    def Banner(self):
        os.system('cls')
        Banner=Banner_font[random.randint(0,13)]
        bnr=pyfiglet.figlet_format('GetLike',font=Banner)
        print()
        print(bnr)
        print('insta:CodeTime_ir                                            https://github.com/NavidAsd')
        print('-------------------------------------"INSTAGRAM"----------------------------------------')
        print()
        print('\33[32m'+'1)Start                                              2)Show All accounts ')
        print()
        print()
        print('3)Log                                                4)Show Proxy list')
        print()
        print()
        print('5)open Config file                                   0)Exit')
        print()
        Bot.Main('00')

    def Main(self,Choice):
        if Choice == '00':
            self.Choice=input('\33[32m'+'Enter Your Choice: '+'\33[37m')
        else:
            self.Choice = Choice

        if self.Choice == '1' or self.Choice=='start' or self.Choice=='Start':
            Bot.Initial_Configuration()
        elif self.Choice == '2':
            subprocess.run(['notepad.exe', './Files/Accounts.txt'])
            Bot.Banner()
        elif self.Choice == '3':
            subprocess.run(['notepad.exe', './Files/log.txt'])
        elif self.Choice == '4':
            subprocess.run(['notepad.exe', './Files/Proxies.txt'])
            Bot.Banner()
        elif self.Choice == '5':
            subprocess.run(['notepad.exe','./Files/Configuration/Config.txt'])
            Bot.Banner()
        elif self.Choice == 'cls':
            Bot.Banner()
        elif self.Choice == '0' or self.Choice == 'Exit' or self.Choice == 'exit' or self.Choice == '-e':
            Bot.Exit()
        else:
            print('\33[31m'+f'"{self.Choice}" is not recognized as an internal or external command.'+'\33[32m')
            print()
            Bot.Main('00')

    def Act_State(self):
        print()
        print('\33[37m'+f'Purchase date: "{self.result["Purchase"]}"  Expiration date: "{self.result["Expiration"]}" Start date: "{self.result["Start"]} "Total number of days: "{self.result["TotalNumber"]}"')
        if (int(self.result["TrialDay"]) >=3):
            print('\33[36m'+f'Yor trial expires in {self.result["TrialDay"]} days.'+'\33[37m')
        else:
            print('\33[31m'+f'Yor trial expires in {self.result["TrialDay"]} days.'+'\33[37m')
            
        print()
        input('\33[36m'+'Press any key to continue..'+'\33[32m')

    def Check_Code(self):
        print('\33[37m'+'"-e", "exit" command for exit')
        self.Code=input('\33[32m'+'Enter License Code: '+'\33[37m')
        if self.Code != "-e" and self.Code != 'exit' and self.Code != 'Exit':
            cv=1
            try:
                for i in range(0,len(self.Code)):
                    if i %4==0  and i!=0:
                        if i==4 and self.Code[i] != '-':
                            self.Code=self.Code[:i] + '-' + self.Code[i:]
                        elif i !=4 and self.Code[i+cv] == '-':
                            cv+=1
                        elif i!=4 and self.Code[i+cv] != '-':
                            self.Code=self.Code[:i+cv] + '-' + self.Code[i+cv:]
                            cv+=1
            except:
                pass
            if len(self.Code) >= 34 and len(self.Code) <= 64:
                return True
            elif len(self.Code) < 34:
                return 'low'
            elif len(self.Code) > 64:
                return 'high'
        else:
            self.Code=''
            Bot.Banner()

    def Get_AllAcountCount(self):
        try:
            file=open("./Files/Accounts.txt","r")
            data=csv.reader(file,delimiter=':')
        except:
            print('\33[31m'+'"Accounts.txt" file not found in "./Files/" path!'+'\33[37m')
            inp=input('\33[32m'+"Create File in Path (y/n)?"+'\33[37m')
            if(inp=='y' or inp=='Y' or inp=='yes' or inp=='Yes' or inp == 'yy'):
                fl=open("./Files/Accounts.txt","w")
                fl.write('GetlikeUser:GetlikePass:InstaUser:InstaPass')
                fl.write('\n')
                fl.write('GetlikeUser:GetlikePass:InstaUser:InstaPass')
                fl.close()
                print('\33[32m'+'After entering the data and closing the Notepad, you can use the program.'+'\33[37m')
                subprocess.run(['notepad.exe', './Files/Accounts.txt'])
            else:
                Bot.Exit()

        for i in data:
            if(("".join(i[:1]).strip()) != '' and ("".join(i[-1]).strip()) != '' and ("".join(i[:1]).strip()) != ("".join(i[-1]).strip()) and ("".join(i[1:2]).strip()) !='' and ("".join(i[2:3]).strip()) != ''):
                self.AccountCount+=1
        Bot.Get_AcountData()

    def Get_AcountData(self):
        try:
            file=open("./Files/Accounts.txt","r")
            data=csv.reader(file,delimiter=':')
        except:
            print('\33[31m'+'"Accounts.txt" file not found in "./Files/" path!'+'\33[37m')
            inp=input('\33[32m'+"Create File in Path (y/n)?"+'\33[37m')
            if(inp=='y' or inp=='Y' or inp=='yes' or inp=='Yes' or inp == 'yy'):
                fl=open("./Files/Accounts.txt","w")
                fl.write('GetlikeUser:GetlikePass:InstaUser:InstaPass')
                fl.write('\n')
                fl.write('GetlikeUser:GetlikePass:InstaUser:InstaPass')
                fl.close()
                print('\33[32m'+'After entering the data and closing the Notepad, you can use the program.'+'\33[37m')
                subprocess.run(['notepad.exe', './Files/Accounts.txt'])
            else:
                Bot.Exit()
        cv=0
        for i in data:
            cv+=1
            if(cv>=self.count and self.count <= self.AccountCount):
                cv=1
                self.User=("".join(i[:1]).strip())
                self.Pass=("".join(i[1:2]).strip())
                self.InstaUser=("".join(i[2:3]).strip())
                self.InstaPass=("".join(i[-1]).strip())
                break
            elif(self.count > self.AccountCount):
                ine=input('\33[32m'+'All used accounts Process to restart (y/n)?'+'\33[37m')
                if(ine == 'y' or ine == 'Y' or ine =='yes' or ine =='Yse' or ine =='yy'):
                    self.AccountCount=0
                    Bot.Initial_Configuration()
                else:
                    Bot.Exit()

        if(self.User != '' and self.Pass != '' and self.User != self.Pass and self.InstaUser!='' and self.InstaPass != '' and self.InstaUser != self.InstaPass):
            Bot.GetStatistics()
        elif(self.User =='' or self.Pass ==''):
            print('\33[31m'+f'|Getlike| Your username: "{self.User}" or password: "{self.Pass}" is Empty or !'+'\33[37m')
            Bot.Change_Account()
        elif(self.InstaUser == '' or self.InstaPass == ''):
            print('\33[31m'+f'|Instagram| Your username: "{self.InstaUser}" or password: "{self.InstaPass}" is Empty or !'+'\33[37m')
            Bot.Change_Account()
        elif(self.User == self.Pass):
            print('\33[31m'+f'|Getlike| Your username: "{self.User}" is the same as your password!'+'\33[37m')
            Bot.Change_Account()
        else:
            print('\33[31m'+f'|Instagram| Your username: "{self.InstaUser}" is the same as your password!'+'\33[37m')
            Bot.Change_Account()

    def Change_Account(self):
        if(self.count <= self.AccountCount):
            Bot.Get_AcountData()
        else:
            inp= input('\33[32m'+'Found a new account Are you making a change(y/n)? '+'\33[37m')
            if(inp == 'y' or inp == 'Y' or inp =='yes' or inp =='Yse' or inp =='yy'):
                subprocess.run(['notepad.exe', './Files/Accounts.txt'])
                self.count = 0
                Bot.Initial_Configuration()
            else:
                Bot.Exit()

    def Start(self):
        if(self.UseChrome):
            try:
                self.option = webdriver.ChromeOptions()
                # For older ChromeDriver under version 79.0.3945.16
                self.option.add_experimental_option("excludeSwitches", ["enable-automation"])
                self.option.add_experimental_option('useAutomationExtension', False)
                self.option.add_argument("window-size=1280,800")
                #..#
                self.driver=webdriver.Chrome(executable_path="./Driver/_ChromeDr_.exe",options=self.option)
                self.driver.set_window_size(width=self.width+1,height=self.height+1)
                self.driver.set_window_position(self.width-15,0)
                pickle.dump( self.driver.get_cookies() , open("./Driver/Cookies.pkl","wb"))
            except (Exception) as ex:
                print('\33[31m'+f'{ex}')
                Bot.Exit()
        else:
            try:
                self.option = webdriver.FirefoxOptions()
                self.option.add_argument("window-size=1280,800")
                self.driver = webdriver.Firefox(executable_path='./Driver/_FireFoxDr_.exe',options=self.option)
                self.driver.set_window_size(width=self.width+1,height=self.height+1)
                self.driver.set_window_position(self.width-15,0)
            except (Exception) as ex:
                print('\33[31m'+f'{ex}')
                Bot.Exit()

        Bot.Login_ToInstagram()   

    def Login_ToInstagram(self):
        self.driver.get(InstagramUrl)
        try:
            time.sleep(0.5)
            self.driver.find_element_by_xpath('/html/body/div[2]/div/div/div/div[2]/button[1]').click()
        except:
            pass

        try:
            time.sleep(1)
            self.WaitForObject(By.XPATH,'//*[@id="loginForm"]/div/div[1]/div/label/input').send_keys(self.InstaUser)
            self.WaitForObject(By.XPATH,'//*[@id="loginForm"]/div/div[2]/div/label/input').send_keys(self.InstaPass,Keys.RETURN)
        except Exception as EX:
            print(EX)

        time.sleep(0.1)
        try:
            self.WaitForObject(By.XPATH,'//*[@id="slfErrorAlert"]')
            print()
            print('\33[31m'+'We Detected An Unusual Login Attempt!')
            #print('\33[32m'+'Please fix the problem and run the program again.')
            print('\33[32m'+'Change Account'+'\33[37m')
            #time.sleep(2)
            #Bot.Exit()
            self.driver.quit()
            self.count += 1
            Bot.ChangeAccount()
        except:
            pass

        try:
            #not now button cookie
            self.WaitForObject(By.CSS_SELECTOR,'#react-root > section > main > div > div > div > div > button').click()
        except:
            pass

        try:
            time.sleep(0.4)
            #not now button notification
            self.driver.find_element_by_css_selector('body > div.RnEpo.Yx5HN > div > div > div > div.mt3GC > button.aOOlW.HoLwm').click()
        except:
            pass

        try:
            self.driver.find_element_by_css_selector('._0ZPOP > svg:nth-child(1)')
            print('\33[32m'+"Login to Instagram Was Successfully"+'\33[37m')
        except:
            print('\33[31m'+'login to instagram failed try again'+'\33[37m')
            Bot.Login_ToInstagram()
        
        Bot.LoginTo_Getlike()

    def LoginTo_Getlike(self):
        self.driver.execute_script(f"window.open('{GetlikeUrl}');")
        window_name = self.driver.window_handles[1]
        self.driver.switch_to.window(window_name=window_name)
        self.WaitForObject(By.CSS_SELECTOR,'#User_loginLogin').send_keys(self.User)
        self.WaitForObject(By.CSS_SELECTOR,'#User_passwordLogin').send_keys(self.Pass)
        Bot.Recaptcha()  

    def Recaptcha(self):
        try:
            self.driver.switch_to_frame(self.driver.find_element_by_tag_name("iframe"))
            self.WaitForObject(By.CLASS_NAME,'recaptcha-checkbox goog-inline-block recaptcha-checkbox-unchecked rc-anchor-checkbox').click()
            time.sleep(1)
            Bot.Alarm()
        except:
            print('\33[31m'+'Captcha Not Found!'+'\33[37m')
            try:
                self.driver.find_element_by_css_selector('input.btn:nth-child(4)').click()
                time.sleep(0.3)
            except:
                pass
            Bot.CheckState()

    def Alarm(self):
        call(["osascript -e 'set volume output volume 100'"], shell=True)
        try:
            playsound.playsound('./Files/Sound/Alarm.wav')
        except:
            print('\33[32m'+'Could not find "Alarm.wav" file in "./Files/Sound/" '+'\33[37m')
        try:
            playsound.playsound('./Files/Sound/Voice.wav')
        except:
            print('\33[32m'+'Could not find "Voice.wav" file in "./Files/Sound/" '+'\33[37m')
        print()
        print('\33[32m'+'Please resolve the login page captcha..'+'\33[37m')
        Bot.CheckState()

    def CheckState(self):
        try:
            #check login 
            self.driver.find_element_by_css_selector('#tasks-page > div.page.page-fixed > nav > div > div.navbar-collapse.hidden-xs > ul:nth-child(1) > li.active > a > span.nav-item-icon')
            self.ccv=0
        except:
            try:
                Erortext=self.driver.find_element_by_css_selector('#formLogin > div.alert.alert-danger.error-summary > ul > li').text
                if(Erortext=='Неверный логин или пароль'):
                    print('\33[32m'+f'UserName: "{self.User}" or Password Incorrect!'+'\33[37m')
                    self.count+=1
                    Bot.Get_AcountData()
                elif(Erortext=='Неверная капча'):
                    print()
                    print('\33[32m'+'Please resolve the login page captcha..'+'\33[37m')
            except:
                pass
            self.ccv+=1
            if(self.ccv %2 !=0):
                self.alarmcount+=1
                print('\33[31m'+f'Repeat the alarm [{self.alarmcount}]'+'\33[37m')
                Bot.Alarm()               
            Bot.CheckState()
        try:
            #folow page
            self.driver.get(FolowUrl)
            self.driver.maximize_window()
            time.sleep(0.5)
            self.FolowCount= self.WaitForObject(By.CSS_SELECTOR,'#list-3-3 > span.badge.badge-square.badge-primary.pull-right').text
            self.LikeCount= self.WaitForObject(By.CSS_SELECTOR,'#list-3-1 > span.badge.badge-square.badge-primary.pull-right').text
            self.driver.set_window_size(width=self.width+1,height=self.height+1)
            self.driver.set_window_position(self.width-15,0)
        except:
            pass
        self.FolowCount=int(self.FolowCount)
        self.LikeCount=int(self.LikeCount)
        
        Bot.Clculate_Tasks()
        Bot.Followings()
        #self.driver.get(LikeUrl)
        #Bot.Like()

    def Clculate_Tasks(self):
        if(self.FolowCount > 20):
            self.RepeatFolowTask= self.FolowCount / 20
            self.RepeatFolowTask= math.ceil(self.RepeatFolowTask) +8
        else:
            self.RepeatFolowTask= 3

        if(self.LikeCount > 20):
            self.RepeatLikeTask= self.LikeCount / 20
            self.RepeatLikeTask= math.ceil(self.RepeatLikeTask) +8
        else:
            self.RepeatLikeTask= 3

    def Followings(self):
        Bot.CheckAccountLimit()
        if(self.FolloAccess):
            if(self.Folowed <= self.FolowCount):
                try:
                    try:
                        self.driver.switch_to.window(self.driver.window_handles[2])
                        self.driver.close()
                    except:
                        pass
                    #click Folow Button in getlike.io
                    time.sleep(0.20)
                    task = self.WaitForObject(By.CLASS_NAME,'do-task')
                    Bot.CheckTaskFailed(task.get_attribute('data-task-id'))
                    task.click()
                except:
                    try:
                        self.driver.switch_to.window(self.driver.window_handles[1])
                        #click on Folow Button in Getlike.io again
                        task = self.driver.find_element_by_xpath('//*[@data-do-class="do-task"]')
                        Bot.CheckTaskFailed(task.get_attribute('data-task-id'))
                        task.click()
                    except:
                        try:
                            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                            self.WaitForObject(By.CLASS_NAME,'do.do-task.btn.btn-sm.btn-primary.btn-block').click()
                        except:
                            if(self.RepeatFolowTask >= 0):
                                try:
                                    self.driver.switch_to.window(self.driver.window_handles[1])
                                    #tasks more
                                    #self.driver.get(FolowUrl)
                                    self.WaitForObject(By.CSS_SELECTOR,'#task_more > button').click()
                                    self.driver.execute_script("window.scrollTo(10,document.body.scrollUp)")
                                except:
                                    print('\33[34m'+'Follow tasks completed.'+'\33[37m')
                                    try:
                                        self.driver.maximize_window()
                                        time.sleep(0.5)
                                        money=self.driver.find_element_by_css_selector('#user_money_balance').text
                                        print('\33[32m'+f'Your Money: {money}'+'\33[37m')
                                        self.driver.set_window_size(width=self.width+1,height=self.height+1)
                                        self.driver.set_window_position(self.width-15,0)
                                    except:
                                        pass
                                    self.driver.get(LikeUrl)
                                    Bot.Like()
                                time.sleep(2)
                                self.RepeatFolowTask -=1
                                Bot.Followings() 
                            else:
                                print('\33[34m'+'Follow tasks completed.'+'\33[37m')
                                try:
                                    self.driver.maximize_window()
                                    time.sleep(0.5)
                                    money=self.driver.find_element_by_css_selector('#user_money_balance').text
                                    print('\33[32m'+f'Your Money: {money}'+'\33[37m')
                                    self.driver.set_window_size(width=self.width+1,height=self.height+1)
                                    self.driver.set_window_position(self.width-15,0)
                                except:
                                    pass
                                self.driver.get(LikeUrl)
                                Bot.Like()
                try:
                    time.sleep(0.3)
                    self.driver.switch_to.window(self.driver.window_handles[2])
                except:
                    pass
                try:
                    time.sleep(5)
                    pickle.dump( self.driver.get_cookies() , open("./Driver/Cookies.pkl","wb"))
                    self.driver.switch_to.window(self.driver.window_handles[2])
                    try:
                        time.sleep(0.6)
                        #check user folowed or not in instagram page 
                        self.driver.find_element_by_css_selector('#react-root > section > main > div > header > section > div.XBGH5 > div.qF0y9.Igw0E.IwRSH.eGOV_.ybXk5._4EzTm.bPdm3 > div > div.qF0y9.Igw0E.IwRSH.eGOV_._4EzTm.soMvl > div > span > span.vBF20._1OSdk > button')
                    except:
                        try:
                            time.sleep(0.1)
                            #click Folow Button in instagram 
                            self.WaitForObject(By.CSS_SELECTOR,'#react-root > section > main > div > header > section > div.XBGH5 > div.qF0y9.Igw0E.IwRSH.eGOV_.ybXk5._4EzTm.bPdm3 > div > div > div > span > span.vBF20._1OSdk > button').click()
                            Bot.FollowPlus(1)
                        except:
                            try:
                                self.WaitForObject(By.CSS_SELECTOR,'#react-root > section > main > div > header > section > div.XBGH5 > div.qF0y9.Igw0E.IwRSH.eGOV_.ybXk5._4EzTm.bPdm3 > div > div > button').click()
                                Bot.FollowPlus(1)
                            except:
                                pass
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[1])
                    self.WaitForObject(By.CLASS_NAME,'do.btn.btn-sm.btn-primary.btn-block.btn-success.check-task').click()
                    time.sleep(2)
                except:
                    pass
                self.Folowed+=1
                Bot.Followings()
            else:
                self.driver.get(LikeUrl)
                Bot.Like()
        elif(self.LikeAccess):
            self.driver.get(LikeUrl)
            Bot.Like()
        else:
            print('\33[32m'+'The program is limited by you, there is no possibility to like and follow'+'\33[37m')
            Bot.Main('5')

    def Like(self):
        Bot.CheckAccountLimit()
        if(self.LikeAccess):
            if(self.Liked <= self.LikeCount):
                try:
                    try:
                        self.driver.switch_to.window(self.driver.window_handles[2])
                        self.driver.close()
                    except:
                        pass
                    self.driver.switch_to.window(self.driver.window_handles[1])
                    #click on Like Button in Getlike.io
                    time.sleep(0.20)
                    task = self.WaitForObject(By.CLASS_NAME,'do-task')
                    Bot.CheckTaskFailed(task.get_attribute('data-task-id'))
                    task.click()
                except:
                    try:
                        self.driver.switch_to.window(self.driver.window_handles[1])
                        #click on Like Button in Getlike.io again
                        task = self.driver.find_element_by_xpath('//*[@data-do-class="do-task"]')
                        Bot.CheckTaskFailed(task.get_attribute('data-task-id'))
                        task.click()
                    except:
                        try:
                            self.driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                            #click on Like Button in Getlike.io again
                            task = self.WaitForObject(By.CLASS_NAME,'do.do-task.btn.btn-sm.btn-primary.btn-block')
                            Bot.CheckTaskFailed(task.get_attribute('data-task-id'))
                            task.click()
                        except:
                            if(self.RepeatLikeTask >= 0):
                                try:
                                    self.driver.switch_to.window(self.driver.window_handles[1])
                                    #tasks more
                                    self.WaitForObject(By.CSS_SELECTOR,'#task_more > button').click()
                                    self.driver.execute_script("window.scrollTo(10,document.body.scrollUp)")
                                except:
                                    print('\33[34m'+'Like tasks completed.'+'\33[37m')
                                    try:
                                        self.driver.maximize_window()
                                        time.sleep(0.5)
                                        money=self.driver.find_element_by_css_selector('#user_money_balance').text
                                        print('\33[32m'+f'Your Money: {money}'+'\33[37m')
                                        self.driver.set_window_size(width=self.width+1,height=self.height+1)
                                        self.driver.set_window_position(self.width-15,0)
                                    except:
                                        pass
                                    Bot.ChangeAccount()
                                    
                                time.sleep(2)
                                self.RepeatLikeTask -=1
                                Bot.Like() 
                            else:
                                if(self.count < self.AccountCount):
                                    print('\33[32m'+'No more tasks found'+'\33[37m')
                                    
                                    try:
                                        self.driver.maximize_window()
                                        time.sleep(0.5)
                                        money=self.driver.find_element_by_css_selector('#user_money_balance').text
                                        print('\33[32m'+f'Your Money: {money}'+'\33[37m')
                                        self.driver.set_window_size(width=self.width+1,height=self.height+1)
                                        self.driver.set_window_position(self.width-15,0)
                                    except:
                                        pass
                                    
                                    Bot.ChangeAccount()
                                else:
                                    print('\33[32m'+'Tasks All Account Completed!')
                                    try:
                                        self.driver.maximize_window()
                                        time.sleep(0.5)
                                        money=self.driver.find_element_by_css_selector('#user_money_balance').text
                                        print('\33[32m'+f'Your Money: {money}'+'\33[37m')
                                        self.driver.set_window_size(width=self.width+1,height=self.height+1)
                                        self.driver.set_window_position(self.width-15,0)
                                    except:
                                        pass
                                    print('\33[34m'+'Good Luck.'+'\33[37m')
                                    Bot.Exit()
                try:
                    time.sleep(0.3)
                    self.driver.switch_to.window(self.driver.window_handles[2])
                except:
                    #self.Liked+=1
                    Bot.Like()
                try:
                    time.sleep(5)
                    self.driver.switch_to.window(self.driver.window_handles[2])
                    try:
                        #check liked or not in instagram page
                        aria_label = self.WaitForObject(By.CSS_SELECTOR,'#react-root > section > main > div > div.ltEKP > article > div > div.qF0y9.Igw0E.IwRSH.eGOV_._4EzTm > div > div.eo2As > section.ltpMr.Slqrh > span.fr66n > button > div.QBdPU.B58H7 > svg').get_attribute('aria-label')
                        if(aria_label == 'Like' or aria_label == 'like'):
                            #click on Like Button in instagram
                            time.sleep(0.1)
                            self.WaitForObject(By.CSS_SELECTOR,'#react-root > section > main > div > div.ltEKP > article > div > div.qF0y9.Igw0E.IwRSH.eGOV_._4EzTm > div > div.eo2As > section.ltpMr.Slqrh > span.fr66n > button').click()
                            Bot.LikePlus(1)
                    except:
                        pass
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[1])
                    self.WaitForObject(By.CLASS_NAME,'do.btn.btn-sm.btn-primary.btn-block.btn-success.check-task').click()
                    time.sleep(2)
                    self.Liked+=1
                    Bot.Like()
                except:
                    self.Liked+=1
                    Bot.Like()

            else:
                print('\33[32m'+'No more tasks found!'+'\33[37m')
                try:
                    self.driver.maximize_window()
                    time.sleep(0.5)
                    money=self.driver.find_element_by_css_selector('#user_money_balance').text
                    print('\33[32m'+f'Your Money: {money}'+'\33[37m')
                    self.driver.set_window_size(width=self.width+1,height=self.height+1)
                    self.driver.set_window_position(self.width-15,0)
                except:
                    pass
                Bot.ChangeAccount()
        else:
            print('\33[32m'+'The program is limited by you, there is no possibility to like'+'\33[37m')
            Bot.ChangeAccount()

    def CheckAccountLimit(self):
        if(self.LimitDate < (self.LimitDate + datetime.timedelta(hours=24))):
            if(self.LikeLimitInDay >= self.LikeLimit):
                self.LikeAccess = False
            if(self.FollowLimitInDay >= self.FollowLimit):
                self.FolloAccess = False 
            if(self.FolloAccess == False and self.LikeAccess == False):
                print('\33[32m'+'The allowed number of likes and followers accounts has expired'+'\33[37m')
                Bot.ChangeAccount()
        else:
            Bot.WriteStatistics(1,1)
            
    def ChangeAccount(self):
        if(self.count < self.AccountCount):
            print('\33[32m'+'Change Account...'+'\33[37m')
            self.count+=1
            self.driver.quit()
            Bot.GetStatistics()
        else:
            print('\33[32m'+'Tasks All Account Completed!')
            print('\33[34m'+'Good Luck.'+'\33[37m')
            Bot.Exit()

    def CheckTaskFailed(self,itemid):
        try:
            # Faile Task Just 2 more time
            time.sleep(0.10)
            if(self.NewTaskId == itemid):
                if(self.TaskFaileCount < 2):
                    self.TaskFaileCount += 1
                else:
                    if(self.MoreTaskFaileCount < 3):
                        # more tasks
                        self.WaitForObject(By.CSS_SELECTOR,'#task_more > button').click()
                        self.driver.execute_script("window.scrollTo(10,document.body.scrollUp)")
                        self.MoreTaskFaileCount += 1
                    else:
                        self.MoreTaskFaileCount = 0
                        self.TaskFaileCount = 0
                        self.driver.get(LikeUrl)
                        Bot.Like()
            else: 
                self.TaskFaileCount = 0
                # not sure for moretask to 0 var
                self.MoreTaskFaileCount = 0
                self.NewTaskId = itemid
        except:
            pass

    def LikePlus (self,plus):
        Bot.PlusLimitStatistics(0,plus)

    def FollowPlus(self,plus):
        Bot.PlusLimitStatistics(plus,0)

    def Exit(self):
        print('\33[32m'+'Exit..'+'\33[37m')
        time.sleep(2)
        try:
            self.driver.quit()
        except:
            pass
        pass
        exit()
     
    def test(self):
        pass


Bot = Program('','',0,0,0,0,0,0,0,0,0,'','',0,0,0,'','',True,'',0,'',0,0,0,0,0,0,0,'',0,0,True)
Bot.Banner()
#Bot.test()
